# ---------------- TRANSACTION MODULE USING EXCEL FILE ---------------- #


# Import library
from openpyxl import load_workbook


# Transaction class
class Transaction:

    # Constructor
    def __init__(self, account_number, amount, transaction_type):
        self.account_number = account_number
        self.amount = amount
        self.transaction_type = transaction_type

    # Display transaction details
    def display_transaction(self):
        print("\n--- Transaction Details ---")
        print("Account Number :", self.account_number)
        print("Transaction Type :", self.transaction_type)
        print("Amount :", self.amount)


# Function to update balance in Excel file
def update_balance(account_number, amount, transaction_type):

    # Load Excel workbook
    workbook = load_workbook("details.xlsx")

    # Select active sheet
    sheet = workbook.active

    # Loop through rows
    for row in range(2, sheet.max_row + 1):

        # Get account number from Excel
        excel_account = str(sheet.cell(row=row, column=1).value)

        # Match account number
        if excel_account == account_number:

            # Get current balance
            current_balance = sheet.cell(row=row, column=7).value

            # Deposit
            if transaction_type == "Deposit":
                new_balance = current_balance + amount

            # Withdraw
            elif transaction_type == "Withdraw":

                # Check sufficient balance
                if amount > current_balance:
                    print("Insufficient Balance!")
                    return

                new_balance = current_balance - amount

            # Update balance in Excel
            sheet.cell(row=row, column=7).value = new_balance

            # Save workbook
            workbook.save("details.xlsx")

            print("\nBalance Updated Successfully!")
            print("Updated Balance :", new_balance)

            return

    # If account not found
    print("Account not found!")


# Transaction menu
def transaction_menu():

    while True:

        print("\n========== Transaction Menu ==========")
        print("1. Deposit")
        print("2. Withdraw")
        print("3. Transaction History")
        print("4. Exit")

        choice = input("Select an option: ")

        # ---------------- DEPOSIT ---------------- #
        if choice == '1':

            account_number = input("Enter account number: ")
            amount = float(input("Enter amount to deposit: "))

            # Create transaction object
            transaction = Transaction(
                account_number,
                amount,
                "Deposit"
            )

            # Update Excel balance
            update_balance(account_number, amount, "Deposit")

            # Display transaction details
            transaction.display_transaction()

        # ---------------- WITHDRAW ---------------- #
        elif choice == '2':

            account_number = input("Enter account number: ")
            amount = float(input("Enter amount to withdraw: "))

            # Create transaction object
            transaction = Transaction(
                account_number,
                amount,
                "Withdraw"
            )

            # Update Excel balance
            update_balance(account_number, amount, "Withdraw")

            # Display transaction details
            transaction.display_transaction()

        # ---------------- TRANSACTION HISTORY ---------------- #
        elif choice == '3':

            account_number = input("Enter account number: ")
            display_transaction_history(account_number)

        # ---------------- EXIT ---------------- #
        elif choice == '4':

            print("Exiting Transaction Module...")
            break

        # ---------------- INVALID OPTION ---------------- #
        else:
            print("Invalid option! Please try again.")


# ---------------- DRIVER CODE ---------------- #
