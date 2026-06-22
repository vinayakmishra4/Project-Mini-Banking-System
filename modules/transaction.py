# ---------------- TRANSACTION MODULE USING EXCEL FILE ---------------- #
import os
from datetime import datetime
from openpyxl import load_workbook

# All Excel data files now live in the data/ folder.
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
DETAILS_FILE = os.path.join(DATA_DIR, "details.xlsx")


# ---------------- TRANSACTION CLASS ---------------- #
class Transaction:
    def __init__(self, account_number, amount, transaction_type):
        self.account_number = account_number
        self.amount = amount
        self.transaction_type = transaction_type

    def display_transaction(self):
        print("\n--- Transaction Details ---")
        print("Account Number :", self.account_number)
        print("Transaction Type :", self.transaction_type)
        print("Amount :", self.amount)


# ---------------- SAVE TRANSACTION HISTORY ---------------- #
def save_transaction(account_number, amount, transaction_type):
    workbook = load_workbook(DETAILS_FILE)

    if "TransactionHistory" not in workbook.sheetnames:
        sheet = workbook.create_sheet("TransactionHistory")
        sheet.append([
            "Account Number",
            "Transaction Type",
            "Amount",
            "Date & Time"
        ])
    else:
        sheet = workbook["TransactionHistory"]

    sheet.append([
        account_number,
        transaction_type,
        amount,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    workbook.save(DETAILS_FILE)


# ---------------- UPDATE BALANCE ---------------- #
def update_balance(account_number, amount, transaction_type):
    workbook = load_workbook(DETAILS_FILE)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        excel_account = str(sheet.cell(row=row, column=1).value)

        if excel_account == account_number:
            current_balance = sheet.cell(row=row, column=7).value

            if transaction_type == "Deposit":
                new_balance = current_balance + amount
            elif transaction_type == "Withdraw":
                if amount > current_balance:
                    print("Insufficient Balance!")
                    return False
                new_balance = current_balance - amount

            sheet.cell(row=row, column=7).value = new_balance
            workbook.save(DETAILS_FILE)

            print("\nBalance Updated Successfully!")
            print("Updated Balance :", new_balance)
            return True

    print("Account not found!")
    return False


# ---------------- VIEW TRANSACTION HISTORY ---------------- #
def view_transaction_history(account_number):
    # NOTE: originally this read from a different file ("transaction_history.xlsx")
    # than save_transaction wrote to ("details.xlsx"), so history lookups always
    # came back empty. Fixed to use the same DETAILS_FILE as save_transaction.
    if not os.path.exists(DETAILS_FILE):
        print("\nNo transaction history found.")
        return

    workbook = load_workbook(DETAILS_FILE)

    if "TransactionHistory" not in workbook.sheetnames:
        print("\nNo transaction history found.")
        return

    sheet = workbook["TransactionHistory"]
    found = False

    print("\n========== TRANSACTION HISTORY ==========")
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == account_number:
            found = True
            transaction_type = sheet.cell(row=row, column=2).value
            amount = sheet.cell(row=row, column=3).value
            date_time = sheet.cell(row=row, column=4).value
            print(
                f"Type: {transaction_type} | "
                f"Amount: {amount} | "
                f"Date: {date_time}"
            )

    if not found:
        print("No transactions found for this account.")


# ---------------- TRANSACTION MENU ---------------- #
def transaction_menu():
    while True:
        print("\n========== Transaction Menu ==========")
        print("1. Deposit")
        print("2. Withdraw")
        print("3. Transaction History")
        print("4. Exit")

        choice = input("Select an option: ")

        # ---------------- DEPOSIT ---------------- #
        if choice == '1':
            account_number = input("Enter account number: ")
            amount = float(input("Enter amount to deposit: "))

            transaction = Transaction(account_number, amount, "Deposit")
            success = update_balance(account_number, amount, "Deposit")

            if success:
                save_transaction(account_number, amount, "Deposit")
                transaction.display_transaction()

        # ---------------- WITHDRAW ---------------- #
        elif choice == '2':
            account_number = input("Enter account number: ")
            amount = float(input("Enter amount to withdraw: "))

            transaction = Transaction(account_number, amount, "Withdraw")
            success = update_balance(account_number, amount, "Withdraw")

            if success:
                save_transaction(account_number, amount, "Withdraw")
                transaction.display_transaction()

        # ---------------- TRANSACTION HISTORY ---------------- #
        elif choice == '3':
            account_number = input("Enter account number: ")
            view_transaction_history(account_number)

        # ---------------- EXIT ---------------- #
        elif choice == '4':
            print("Exiting Transaction Module...")
            break

        # ---------------- INVALID OPTION ---------------- #
        else:
            print("Invalid option! Please try again.")
